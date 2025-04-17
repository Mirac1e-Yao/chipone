import openpyxl
from openpyxl import load_workbook

def extract_sheets_to_txt(input_excel, output_txt, start_sheet, end_sheet):
    wb = load_workbook(input_excel, data_only=True)
    
    # 将用户输入的1-based索引转换为0-based
    start_idx = start_sheet - 1
    end_idx = end_sheet - 1
    
    with open(output_txt, 'w', encoding='utf-8') as f:
        # 遍历指定范围内的sheet
        for sheet in wb.worksheets[start_idx:end_idx+1]:
            line = "\nR9F " + f"{(wb.worksheets.index(sheet) - 3):02X}" + "\n"
            f.write(line)
            # 存储B列合并区域信息 {起始行: (结束行, 合并值)}
            merged_dict = {}
            
            # 收集当前sheet的B列合并单元格信息
            for merged_range in sheet.merged_cells.ranges:
                if merged_range.min_col == 2 and merged_range.max_col == 2:
                    min_row = merged_range.min_row
                    max_row = merged_range.max_row
                    merged_value = sheet.cell(min_row, 2).value
                    merged_dict[min_row] = (max_row, merged_value)
            
            current_row = 1
            max_row_in_sheet = sheet.max_row
            
            # 使用while循环控制行号
            while current_row <= max_row_in_sheet:
                # 检查当前行是否在合并区域起始行
                if current_row in merged_dict:
                    max_r, b_value = merged_dict[current_row]
                    
                    # 检查合并区域中是否存在NORMAL
                    has_normal = any(
                        sheet.cell(r, 14).value == 'NORMAL' 
                        for r in range(current_row, max_r + 1)
                    )
                    
                    if has_normal:
                        # 提取合并区域所有M列值
                        m_values = [
                            str(sheet.cell(r, 13).value or '').replace('h', '')  # 删除 'h' 字符
                            for r in range(current_row, max_r + 1)
                        ]
                        line = f"R{b_value} " + " ".join(m_values) + "\n"
                        f.write(line)
                    
                    # 直接跳转到合并区域结束行之后
                    current_row = max_r + 1
                else:
                    # 处理单行数据
                    n_value = sheet.cell(current_row, 14).value
                    if n_value == 'NORMAL':
                        b_value = sheet.cell(current_row, 2).value or ''
                        m_value = str(sheet.cell(current_row, 13).value or '').replace('h', '')  # 删除 'h' 字符
                        line = f"R{b_value} {m_value}\n"
                        f.write(line)
                    current_row += 1
    
    wb.close()

# 使用示例
input_excel = "C:\\Users\\67064\\Desktop\\ICNA3508A_Instruction_Table_OTP.xlsx"
output_txt = "C:\\Users\\67064\\Desktop\\output_3508A_default.txt"
start_sheet = 5   # 起始sheet编号（基于1的索引）
end_sheet = 20     # 结束sheet编号

extract_sheets_to_txt(input_excel, output_txt, start_sheet, end_sheet)